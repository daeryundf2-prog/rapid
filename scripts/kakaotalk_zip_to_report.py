#!/usr/bin/env python3
from __future__ import annotations

# Force UTF-8 stdio so JSON output with non-ASCII evidence text (e.g.
# Korean filenames) survives Windows consoles whose default codec is cp1252.
import sys as _sys

if hasattr(_sys.stdout, "reconfigure"):
    _sys.stdout.reconfigure(encoding="utf-8")
if hasattr(_sys.stderr, "reconfigure"):
    _sys.stderr.reconfigure(encoding="utf-8")

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import shutil
import sqlite3
import sys
import tempfile
from pathlib import Path
from typing import Mapping, Sequence

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from rapidtriage.core.kakaotalk import (  # noqa: E402
    KakaoTalkDecryptError,
    build_kakaotalk_media_inventory,
    extract_zip_archive_safely,
    run_kakaotalk_decrypt,
    run_kakaotalk_sqlcipher_probe,
)


FORBIDDEN_SECRET_KEYS = {
    "key_hex",
    "database_key_hex",
    "object_key",
    "object_key_hex",
    "ikm",
    "ikm_hex",
    "raw_key",
    "raw_key_hex",
    "sqlcipher_raw_key_with_salt_hex",
}

SINGLE_FILE_COPY_LIMIT_BYTES = 512 * 1024 * 1024


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build a redacted PC KakaoTalk data report from a ZIP, extracted folder, or evidence file."
    )
    parser.add_argument("source", help="KakaoTalk ZIP, extracted folder, NTUSER.DAT, DMG, or other evidence file")
    parser.add_argument("--output-dir", required=True, help="Directory for JSON/CSV/XLSX report outputs")
    parser.add_argument("--sqlcipher-bin", default="sqlcipher", help="SQLCipher binary")
    parser.add_argument("--openssl-bin", default="openssl", help="OpenSSL binary for legacy AES-CBC operations")
    parser.add_argument("--timeout-seconds", type=float, default=5.0)
    parser.add_argument("--max-message-residues", type=int, default=1000)
    parser.add_argument("--keep-opened-sqlite", action="store_true", help="Keep plaintext SQLite exports next to the report")
    parser.add_argument(
        "--analysis-mode",
        choices=("both", "auto", "postpatch", "legacy"),
        default="both",
        help="Which KakaoTalk generations to try. both tries post-patch SQLCipher and legacy EDB decrypt.",
    )
    parser.add_argument(
        "--legacy-user-id",
        default=None,
        help="Authorized legacy PC KakaoTalk user id. If omitted, RAPIDTRIAGE_KAKAO_USER_ID is still honored.",
    )
    parser.add_argument(
        "--legacy-pragma",
        default=None,
        help="Authorized legacy PRAGMA key material. Redacted from outputs; RAPIDTRIAGE_KAKAO_PRAGMA is also honored.",
    )
    parser.add_argument(
        "--disable-legacy-fallback",
        action="store_true",
        help="Do not try the legacy pre-BigBang EDB decrypt fallback when the SQLCipher path finds no messages.",
    )
    parser.add_argument("--no-xlsx", action="store_true", help="Skip XLSX generation")
    args = parser.parse_args()

    source = Path(args.source).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    opened_dir = output_dir / "opened_sqlite" if args.keep_opened_sqlite else None
    temp_roots: list[tempfile.TemporaryDirectory[str]] = []
    try:
        source_kind = detect_source_kind(source)
        root = resolve_source_root(source, temp_roots)
        export_dir = opened_dir or Path(tempfile.mkdtemp(prefix="rapidtriage-kakao-opened-"))
        if opened_dir:
            opened_dir.mkdir(parents=True, exist_ok=True)
        payload = run_kakaotalk_sqlcipher_probe(
            root,
            output=output_dir / "kakaotalk_probe.json",
            sqlcipher_bin=args.sqlcipher_bin,
            max_message_residues=args.max_message_residues,
            include_message_preview=True,
            timeout_seconds=args.timeout_seconds,
            export_opened_dir=export_dir,
        )
        assert_no_forbidden_secret_keys(payload)
        postpatch_rooms = build_room_rows(payload)
        postpatch_messages = build_message_rows(payload)
        postpatch_media = build_media_rows(payload, analysis_method="postpatch-sqlcipher")
        rooms = list(postpatch_rooms)
        messages = list(postpatch_messages)
        media = list(postpatch_media)
        legacy_payload: Mapping[str, object] | None = None
        legacy_export_dir: Path | None = None
        legacy_export_created = False
        if not args.disable_legacy_fallback and should_try_legacy_analysis(args.analysis_mode, payload, postpatch_messages):
            legacy_export_dir = opened_dir or Path(tempfile.mkdtemp(prefix="rapidtriage-kakao-legacy-opened-"))
            legacy_export_created = opened_dir is None
            if opened_dir:
                opened_dir.mkdir(parents=True, exist_ok=True)
            legacy_payload = run_kakaotalk_decrypt(
                root,
                output=output_dir / "kakaotalk_legacy_decrypt.json",
                pragma=args.legacy_pragma,
                user_id=args.legacy_user_id,
                include_message_preview=True,
                write_decrypted=True,
                decrypted_dir=legacy_export_dir,
                max_messages_per_db=args.max_message_residues,
                openssl_bin=args.openssl_bin,
                postpatch_memory_carve=False,
            )
            assert_no_forbidden_secret_keys(legacy_payload)
            legacy_messages = build_legacy_message_rows(legacy_payload)
            if legacy_messages:
                legacy_rooms = build_legacy_room_rows(legacy_messages)
                legacy_sqlites = legacy_decrypted_sqlite_paths(legacy_payload)
                legacy_media_inventory = build_kakaotalk_media_inventory(
                    root=root,
                    exported_sqlite_paths=legacy_sqlites,
                    include_message_preview=True,
                )
                legacy_media = build_media_rows(
                    {"postpatch_media_inventory": legacy_media_inventory},
                    analysis_method="legacy-edb-decrypt",
                )
                rooms.extend(legacy_rooms)
                messages.extend(legacy_messages)
                media.extend(legacy_media)
        summary = build_summary_payload(
            source=source,
            payload=payload,
            legacy_payload=legacy_payload,
            rooms=rooms,
            messages=messages,
            media=media,
            temporary_extraction=source.is_file() and source.suffix.lower() == ".zip",
            source_kind=source_kind,
            input_note=input_note_for_source(source, source_kind),
        )
        xlsx_created = False
        if not args.no_xlsx:
            xlsx_created = write_xlsx(
                output_dir / "kakaotalk_report.xlsx",
                summary=summary,
                rooms=rooms,
                messages=messages,
                media=media,
            )
        if not xlsx_created:
            summary["outputs"].pop("xlsx", None)
            summary["xlsx_created"] = False
        else:
            summary["xlsx_created"] = True
        report_payload = {
            "summary": summary,
            "postpatch_probe": payload,
            "legacy_decrypt": legacy_payload,
        }
        assert_no_forbidden_secret_keys(report_payload)
        write_json(output_dir / "kakaotalk_report.json", report_payload)
        write_json(output_dir / "kakaotalk_summary.json", summary)
        write_csv(output_dir / "kakaotalk_rooms.csv", rooms)
        write_csv(output_dir / "kakaotalk_messages.csv", messages)
        write_csv(output_dir / "kakaotalk_media.csv", media)
        write_csv(output_dir / "kakaotalk_database_counts.csv", build_database_count_rows(payload, legacy_payload))
        if not args.keep_opened_sqlite and export_dir.exists():
            shutil.rmtree(export_dir, ignore_errors=True)
        if legacy_export_created and legacy_export_dir and legacy_export_dir.exists():
            shutil.rmtree(legacy_export_dir, ignore_errors=True)
    finally:
        for temp_root in temp_roots:
            temp_root.cleanup()

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def detect_source_kind(source: Path) -> str:
    if source.is_dir():
        return "directory"
    suffix = source.suffix.lower()
    if suffix == ".zip":
        return "zip"
    if source.name.lower() == "ntuser.dat":
        return "ntuser-dat"
    if suffix == ".dmg":
        return "dmg"
    if suffix:
        return f"file:{suffix[1:]}"
    return "file:unknown"


def input_note_for_source(source: Path, source_kind: str) -> str:
    if source_kind == "zip":
        return "ZIP was safely extracted and analyzed."
    if source_kind == "directory":
        return "Folder was analyzed directly."
    if source_kind == "ntuser-dat":
        return "Single NTUSER.DAT input was accepted for registry/device clues; add Kakao DB files or a Kakao folder for conversations."
    if source_kind == "dmg":
        return "DMG was accepted as an evidence file. On Windows, mount/extract the DMG first and analyze the extracted Kakao folder for conversations."
    suffix = source.suffix.lower()
    if suffix in {".alz", ".egg", ".rar", ".7z", ".tar", ".gz"}:
        return "Archive container was accepted as evidence metadata. Extract it first or provide ZIP/folder input for conversation parsing."
    if source.is_file():
        return "Single evidence file was accepted; conversation recovery requires Kakao DB files, registry hives, or an extracted Kakao folder."
    return ""


def resolve_source_root(source: Path, temp_roots: list[tempfile.TemporaryDirectory[str]]) -> Path:
    if source.is_dir():
        return source
    if not source.is_file():
        raise KakaoTalkDecryptError("source must exist and be a KakaoTalk ZIP, extracted folder, or evidence file")
    temp_root = tempfile.TemporaryDirectory(prefix="rapidtriage-kakao-source-report-")
    temp_roots.append(temp_root)
    root = Path(temp_root.name)
    if source.suffix.lower() == ".zip":
        return extract_zip_archive_safely(source, root)
    stage_single_evidence_file(source, root / source.name)
    return root


def stage_single_evidence_file(source: Path, target: Path) -> None:
    try:
        os.link(source, target)
        return
    except OSError:
        pass
    size = source.stat().st_size
    if size <= SINGLE_FILE_COPY_LIMIT_BYTES:
        shutil.copy2(source, target)
        return
    manifest = {
        "source_path": str(source),
        "source_name": source.name,
        "source_size": size,
        "staging": "metadata-only",
        "reason": "single evidence file exceeded safe copy limit",
        "copy_limit_bytes": SINGLE_FILE_COPY_LIMIT_BYTES,
    }
    (target.parent / "single_evidence_file_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_room_rows(payload: Mapping[str, object]) -> list[dict[str, object]]:
    rows = []
    for room in payload.get("postpatch_chat_room_previews", []) or []:
        if not isinstance(room, Mapping):
            continue
        rows.append(
            {
                "chat_id": room.get("chat_id", ""),
                "room_title": room.get("room_title", ""),
                "room_type": room.get("room_type", ""),
                "active_members_count": room.get("active_members_count", ""),
                "last_updated_at": room.get("last_updated_at", ""),
                "last_log_id": room.get("last_log_id", ""),
                "message_text": room.get("message_text", ""),
                "message_text_sha256": room.get("message_text_sha256", ""),
                "source_export_path": room.get("source_export_path", ""),
                "analysis_method": "postpatch-sqlcipher",
            }
        )
    return rows


def build_message_rows(payload: Mapping[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    room_titles = {
        str(room.get("chat_id") or ""): str(room.get("room_title") or "")
        for room in payload.get("postpatch_chat_room_previews", []) or []
        if isinstance(room, Mapping)
    }
    for match in payload.get("matches", []) or []:
        if not isinstance(match, Mapping):
            continue
        export_info = match.get("export")
        if not isinstance(export_info, Mapping) or not export_info.get("exported"):
            continue
        export_path = Path(str(export_info.get("export_path") or ""))
        if not export_path.exists() or "chatLogs_" not in export_path.name:
            continue
        chat_id = chat_id_from_name(export_path.name)
        rows.extend(
            read_chatlog_messages(
                export_path,
                chat_id=chat_id,
                room_title=room_titles.get(chat_id, ""),
                analysis_method="postpatch-sqlcipher",
            )
        )
    return rows


def build_legacy_message_rows(payload: Mapping[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sqlite_path in legacy_decrypted_sqlite_paths(payload):
        chat_id = chat_id_from_name(sqlite_path.name)
        rows.extend(
            read_chatlog_messages(
                sqlite_path,
                chat_id=chat_id,
                room_title="",
                analysis_method="legacy-edb-decrypt",
            )
        )
    return rows


def build_database_count_rows(
    postpatch_payload: Mapping[str, object],
    legacy_payload: Mapping[str, object] | None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    legacy_entries = legacy_payload.get("entries", []) if isinstance(legacy_payload, Mapping) else []
    for entry in legacy_entries or []:
        if not isinstance(entry, Mapping):
            continue
        source_path = str(entry.get("source_path") or "")
        rows.append(
            {
                "analysis_method": "legacy-edb-decrypt",
                "chat_id": str(entry.get("chat_id") or chat_id_from_name(source_path)),
                "source_path": source_path,
                "sqlite_status": str(entry.get("sqlite_status") or ""),
                "decrypt_status": str(entry.get("decrypt_status") or ""),
                "message_row_count": entry.get("message_row_count") or 0,
                "message_preview_count": len(entry.get("message_previews") or []),
                "decrypted_path": str(entry.get("decrypted_path") or ""),
                "source_size": entry.get("source_size") or "",
                "matched_key_derivation": str(entry.get("matched_key_derivation") or ""),
                "matched_key_source": str(entry.get("matched_key_source") or ""),
            }
        )
    if rows:
        return rows
    for match in postpatch_payload.get("matches", []) or []:
        if not isinstance(match, Mapping):
            continue
        export_info = match.get("export")
        if not isinstance(export_info, Mapping):
            export_info = {}
        source_path = str(match.get("source_path") or export_info.get("export_path") or "")
        rows.append(
            {
                "analysis_method": "postpatch-sqlcipher",
                "chat_id": chat_id_from_name(source_path),
                "source_path": source_path,
                "sqlite_status": "opened" if export_info.get("exported") else str(match.get("status") or ""),
                "decrypt_status": str(match.get("status") or ""),
                "message_row_count": match.get("message_row_count") or 0,
                "message_preview_count": len(match.get("message_previews") or []),
                "decrypted_path": str(export_info.get("export_path") or ""),
                "source_size": match.get("source_size") or "",
                "matched_key_derivation": "",
                "matched_key_source": "",
            }
        )
    return rows


def count_database_statuses(payload: Mapping[str, object] | None) -> dict[str, int]:
    counts: dict[str, int] = {}
    if not isinstance(payload, Mapping):
        return counts
    for entry in payload.get("entries", []) or []:
        if not isinstance(entry, Mapping):
            continue
        status = str(entry.get("sqlite_status") or entry.get("decrypt_status") or "unknown")
        counts[status] = counts.get(status, 0) + 1
    return counts


def legacy_decrypted_sqlite_paths(payload: Mapping[str, object]) -> list[Path]:
    paths: list[Path] = []
    for entry in payload.get("entries", []) or []:
        if not isinstance(entry, Mapping):
            continue
        if entry.get("sqlite_status") != "opened":
            continue
        decrypted_path = entry.get("decrypted_path")
        if not decrypted_path:
            continue
        path = Path(str(decrypted_path))
        if path.exists():
            paths.append(path)
    return paths


def build_legacy_room_rows(messages: Sequence[Mapping[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for message in messages:
        chat_id = str(message.get("chat_id") or "")
        if not chat_id:
            continue
        room = grouped.setdefault(
            chat_id,
            {
                "chat_id": chat_id,
                "room_title": "",
                "room_type": "legacy_chatlog",
                "active_members_count": "",
                "last_updated_at": "",
                "last_log_id": "",
                "message_text": "",
                "message_text_sha256": "",
                "source_export_path": message.get("source_export_path", ""),
                "message_count": 0,
                "analysis_method": "legacy-edb-decrypt",
            },
        )
        room["message_count"] = int(room["message_count"] or 0) + 1
        room["last_updated_at"] = message.get("send_at", "")
        room["last_log_id"] = message.get("log_id", "")
    return sorted(grouped.values(), key=lambda row: str(row.get("chat_id") or ""))


def read_chatlog_messages(
    path: Path,
    *,
    chat_id: str,
    room_title: str,
    analysis_method: str,
) -> list[dict[str, object]]:
    connection = sqlite3.connect(f"{path.resolve().as_uri()}?mode=ro&immutable=1", uri=True)
    connection.row_factory = sqlite3.Row
    try:
        table = connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='chatLogs' LIMIT 1"
        ).fetchone()
        if table is None:
            return []
        columns = {str(row[1]) for row in connection.execute("PRAGMA table_info(chatLogs)").fetchall()}
        selected = [
            column
            for column in (
                "logId",
                "authorId",
                "type",
                "sendAt",
                "message",
                "attachement",
                "deleted",
                "prevLogId",
                "referer",
                "threadId",
                "scope",
            )
            if column in columns
        ]
        sql = f"SELECT {', '.join(quote_identifier(column) for column in selected)} FROM chatLogs ORDER BY sendAt, logId"
        rows: list[dict[str, object]] = []
        for record in connection.execute(sql):
            attachment = str(record["attachement"] or "") if "attachement" in record.keys() else ""
            message = str(record["message"] or "") if "message" in record.keys() else ""
            rows.append(
                {
                    "chat_id": chat_id,
                    "room_title": room_title,
                    "log_id": record["logId"] if "logId" in record.keys() else "",
                    "author_id": record["authorId"] if "authorId" in record.keys() else "",
                    "type": record["type"] if "type" in record.keys() else "",
                    "send_at": record["sendAt"] if "sendAt" in record.keys() else "",
                    "send_at_utc": unix_to_iso(record["sendAt"]) if "sendAt" in record.keys() else "",
                    "message": message,
                    "message_sha256": hashlib.sha256(message.encode("utf-8", errors="ignore")).hexdigest() if message else "",
                    "attachment_present": bool(attachment and attachment != "{}"),
                    "attachment_json_sha256": hashlib.sha256(attachment.encode("utf-8", errors="replace")).hexdigest()
                    if attachment
                    else "",
                    "deleted": record["deleted"] if "deleted" in record.keys() else "",
                    "prev_log_id": record["prevLogId"] if "prevLogId" in record.keys() else "",
                    "referer": record["referer"] if "referer" in record.keys() else "",
                    "thread_id": record["threadId"] if "threadId" in record.keys() else "",
                    "scope": record["scope"] if "scope" in record.keys() else "",
                    "source_export_path": str(path),
                    "analysis_method": analysis_method,
                }
            )
        return rows
    finally:
        connection.close()


def build_media_rows(payload: Mapping[str, object], *, analysis_method: str) -> list[dict[str, object]]:
    inventory = payload.get("postpatch_media_inventory")
    if not isinstance(inventory, Mapping):
        return []
    rows = []
    for item in inventory.get("attachments", []) or []:
        if not isinstance(item, Mapping):
            continue
        rows.append(
            {
                "chat_id": item.get("chat_id", ""),
                "log_id": item.get("log_id", ""),
                "send_at": item.get("send_at", ""),
                "author_id": item.get("author_id", ""),
                "message_type": item.get("message_type", ""),
                "media_class": item.get("media_class", ""),
                "display_name": item.get("display_name", ""),
                "mime": item.get("mime", ""),
                "declared_size": item.get("declared_size", ""),
                "width": item.get("width", ""),
                "height": item.get("height", ""),
                "duration": item.get("duration", ""),
                "expire": item.get("expire", ""),
                "checksum_count": item.get("checksum_count", ""),
                "checksum_sha256": item.get("checksum_sha256", ""),
                "basename_candidates": json.dumps(item.get("basename_candidates", []), ensure_ascii=False),
                "local_match_count": item.get("local_match_count", ""),
                "local_matches": json.dumps(item.get("local_matches", []), ensure_ascii=False),
                "review_status": item.get("review_status", ""),
                "message_text_sha256": item.get("message_text_sha256", ""),
                "attachment_json_sha256": item.get("attachment_json_sha256", ""),
                "analysis_method": analysis_method,
            }
        )
    return rows


def build_summary_payload(
    *,
    source: Path,
    payload: Mapping[str, object],
    legacy_payload: Mapping[str, object] | None,
    rooms: Sequence[Mapping[str, object]],
    messages: Sequence[Mapping[str, object]],
    media: Sequence[Mapping[str, object]],
    temporary_extraction: bool,
    source_kind: str,
    input_note: str,
) -> dict[str, object]:
    summary = payload.get("summary", {})
    if not isinstance(summary, Mapping):
        summary = {}
    legacy_summary = legacy_payload.get("summary", {}) if isinstance(legacy_payload, Mapping) else {}
    if not isinstance(legacy_summary, Mapping):
        legacy_summary = {}
    status = summary.get("status", "")
    analysis_mode = "postpatch-sqlcipher"
    if legacy_payload is not None:
        if messages:
            status = "matched"
            analysis_mode = "postpatch-sqlcipher+legacy-edb-decrypt"
        else:
            analysis_mode = "postpatch-sqlcipher+legacy-attempted"
    message_method_counts = count_rows_by_method(messages)
    room_method_counts = count_rows_by_method(rooms)
    media_method_counts = count_rows_by_method(media)
    legacy_sqlite_open_count = int(legacy_summary.get("sqlite_open_count") or 0)
    legacy_raw_message_row_count = int(legacy_summary.get("message_row_count") or 0)
    chat_database_count = int(summary.get("chat_database_count") or 0)
    visible_message_count = len(messages)
    unopened_database_count = max(chat_database_count - legacy_sqlite_open_count, 0) if legacy_payload is not None else 0
    return {
        "source": str(source),
        "source_type": source_kind,
        "input_note": input_note,
        "temporary_extraction": temporary_extraction,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "status": status,
        "analysis_mode": analysis_mode,
        "postpatch_status": summary.get("status", ""),
        "legacy_attempted": legacy_payload is not None,
        "legacy_sqlite_open_count": legacy_sqlite_open_count,
        "legacy_message_row_count": legacy_raw_message_row_count,
        "legacy_database_status_counts": count_database_statuses(legacy_payload),
        "unopened_database_count": unopened_database_count,
        "visible_message_count": visible_message_count,
        "raw_recovered_message_row_count": legacy_raw_message_row_count,
        "message_count_note": (
            "raw_recovered_message_row_count is the total row count in opened DBs; "
            "visible_message_count/message_count is what this viewer/CSV currently displays."
        ),
        "postpatch_message_count": message_method_counts.get("postpatch-sqlcipher", 0),
        "legacy_message_count": message_method_counts.get("legacy-edb-decrypt", 0),
        "message_method_counts": message_method_counts,
        "room_method_counts": room_method_counts,
        "media_method_counts": media_method_counts,
        "chat_database_count": chat_database_count,
        "opened_database_count": summary.get("opened_database_count", 0),
        "room_count": len(rooms),
        "message_count": visible_message_count,
        "media_attachment_count": len(media),
        "local_media_file_count": summary.get("postpatch_media_local_file_count", 0),
        "local_media_match_count": summary.get("postpatch_media_local_match_count", 0),
        "sensitive_keys_exported": False,
        "outputs": {
            "probe_json": "kakaotalk_probe.json",
            "legacy_json": "kakaotalk_legacy_decrypt.json" if legacy_payload is not None else "",
            "report_json": "kakaotalk_report.json",
            "summary_json": "kakaotalk_summary.json",
            "rooms_csv": "kakaotalk_rooms.csv",
            "messages_csv": "kakaotalk_messages.csv",
            "media_csv": "kakaotalk_media.csv",
            "database_counts_csv": "kakaotalk_database_counts.csv",
            "xlsx": "kakaotalk_report.xlsx",
        },
    }


def should_try_legacy_analysis(
    analysis_mode: str,
    payload: Mapping[str, object],
    postpatch_messages: Sequence[Mapping[str, object]],
) -> bool:
    if analysis_mode == "postpatch":
        return False
    if analysis_mode == "both" or analysis_mode == "legacy":
        return True
    if postpatch_messages:
        return False
    summary = payload.get("summary", {})
    if not isinstance(summary, Mapping):
        return True
    return int(summary.get("chat_database_count") or 0) > 0


def count_rows_by_method(rows: Sequence[Mapping[str, object]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        method = str(row.get("analysis_method") or "unknown")
        counts[method] = counts.get(method, 0) + 1
    return counts


def assert_no_forbidden_secret_keys(value: object, *, path: str = "$") -> None:
    if isinstance(value, Mapping):
        for key, child in value.items():
            key_text = str(key)
            if key_text in FORBIDDEN_SECRET_KEYS:
                raise KakaoTalkDecryptError(f"Forbidden sensitive key would be exported at {path}.{key_text}")
            assert_no_forbidden_secret_keys(child, path=f"{path}.{key_text}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            assert_no_forbidden_secret_keys(child, path=f"{path}[{index}]")


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    headers = sorted({key for row in rows for key in row.keys()})
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(
    path: Path,
    *,
    summary: Mapping[str, object],
    rooms: Sequence[Mapping[str, object]],
    messages: Sequence[Mapping[str, object]],
    media: Sequence[Mapping[str, object]],
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["metric", "value"])
    for key, value in summary.items():
        sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    append_sheet(workbook, "Rooms", rooms)
    append_sheet(workbook, "Messages", messages)
    append_sheet(workbook, "Media", media)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="334155")
        for column in range(1, min(sheet.max_column, 20) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 22
    workbook.save(path)
    return True


def append_sheet(workbook: object, title: str, rows: Sequence[Mapping[str, object]]) -> None:
    sheet = workbook.create_sheet(title)
    headers = sorted({key for row in rows for key in row.keys()})
    sheet.append(headers)
    for row in rows:
        sheet.append([json.dumps(row.get(key, ""), ensure_ascii=False) if isinstance(row.get(key, ""), (dict, list)) else row.get(key, "") for key in headers])


def chat_id_from_name(name: str) -> str:
    marker = "chatLogs_"
    if marker not in name:
        return ""
    tail = name.split(marker, 1)[1]
    return tail.split(".", 1)[0]


def quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def unix_to_iso(value: object) -> str:
    try:
        timestamp = int(value)
    except (TypeError, ValueError):
        return ""
    if timestamp <= 0:
        return ""
    try:
        return dt.datetime.fromtimestamp(timestamp, tz=dt.timezone.utc).isoformat()
    except (OverflowError, OSError, ValueError):
        return ""


if __name__ == "__main__":
    raise SystemExit(main())
